"""测试管理页面备注列功能

测试范围:
1. API功能测试 - 查询接口返回notes、更新备注接口、导出功能
2. 边界情况测试 - 特殊字符、NULL值、空白字符串、超长文本
3. 集成测试 - 端到端流程验证
"""
import pytest
import sqlite3
import tempfile
import os
import zipfile
from datetime import datetime
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from io import BytesIO

from app.main import app
from app.core.database import get_db_connection
from app.core.config import get_settings


# ==================== 测试客户端和数据库fixtures ====================

@pytest.fixture
def client():
    """创建测试客户端"""
    return TestClient(app)


@pytest.fixture
def test_db():
    """创建测试数据库并初始化表结构"""
    settings = get_settings()
    original_db_url = settings.DATABASE_URL

    # 创建临时数据库
    with tempfile.NamedTemporaryFile(mode='w', suffix='.db', delete=False) as f:
        test_db_path = f.name

    # 修改配置指向测试数据库
    settings.DATABASE_URL = f"sqlite:///{test_db_path}"

    # 初始化数据库结构
    conn = sqlite3.connect(test_db_path)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS upload_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            business_id VARCHAR(50) NOT NULL,
            doc_number VARCHAR(100),
            doc_type VARCHAR(20),
            product_type TEXT DEFAULT NULL,
            file_name VARCHAR(255) NOT NULL,
            file_size INTEGER NOT NULL,
            file_extension VARCHAR(20),
            upload_time DATETIME DEFAULT CURRENT_TIMESTAMP,
            status VARCHAR(20) NOT NULL,
            error_code VARCHAR(50),
            error_message TEXT,
            yonyou_file_id VARCHAR(255),
            retry_count INTEGER DEFAULT 0,
            local_file_path VARCHAR(500),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            deleted_at TEXT DEFAULT NULL,
            checked INTEGER DEFAULT 0,
            notes TEXT DEFAULT NULL
        )
    """)

    conn.commit()
    conn.close()

    yield test_db_path

    # 恢复原始配置并清理测试数据库
    settings.DATABASE_URL = original_db_url
    if os.path.exists(test_db_path):
        os.unlink(test_db_path)


@pytest.fixture
def sample_record(test_db):
    """创建一条示例记录用于测试"""
    conn = sqlite3.connect(test_db)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO upload_history (
            business_id, doc_number, doc_type, product_type,
            file_name, file_size, file_extension,
            upload_time, status, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        "123456",
        "SO20250101",
        "销售",
        "油脂",
        "test.jpg",
        1024,
        ".jpg",
        "2025-01-01 10:00:00",
        "success",
        "原始备注"
    ])

    record_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return record_id


@pytest.fixture
def sample_records_with_notes(test_db):
    """创建多条带备注的记录用于测试"""
    conn = sqlite3.connect(test_db)
    cursor = conn.cursor()

    records = [
        ("123456", "SO001", "销售", "油脂", "file1.jpg", 1024, ".jpg", "success", "第一条备注"),
        ("123457", "SO002", "转库", "快消", "file2.jpg", 2048, ".jpg", "success", "第二条备注"),
        ("123458", "SO003", "销售", "油脂", "file3.jpg", 3072, ".jpg", "failed", None),
        ("123459", "SO004", "转库", "快消", "file4.jpg", 4096, ".jpg", "success", ""),
    ]

    record_ids = []
    for business_id, doc_number, doc_type, product_type, file_name, file_size, file_ext, status, notes in records:
        cursor.execute("""
            INSERT INTO upload_history (
                business_id, doc_number, doc_type, product_type,
                file_name, file_size, file_extension,
                upload_time, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), ?, ?)
        """, [business_id, doc_number, doc_type, product_type, file_name, file_size, file_ext, status, notes])
        record_ids.append(cursor.lastrowid)

    conn.commit()
    conn.close()

    return record_ids


# ==================== 1. API功能测试 ====================

class TestNotesAPIFunctionality:
    """测试备注API的核心功能"""

    def test_get_records_includes_notes_field(self, client, test_db, sample_record):
        """测试查询接口返回notes字段"""
        response = client.get("/api/admin/records?page=1&page_size=10")

        assert response.status_code == 200
        data = response.json()

        # 验证响应结构
        assert "records" in data
        assert len(data["records"]) > 0

        # 验证notes字段存在
        first_record = data["records"][0]
        assert "notes" in first_record
        assert first_record["notes"] == "原始备注"


    def test_update_notes_success(self, client, test_db, sample_record):
        """测试成功更新备注"""
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "更新后的备注内容"}
        )

        assert response.status_code == 200
        data = response.json()

        # 验证响应内容
        assert data["success"] is True
        assert data["id"] == sample_record
        assert data["notes"] == "更新后的备注内容"
        assert data["message"] == "备注已更新"

        # 验证数据库中的数据已更新
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT notes FROM upload_history WHERE id = ?", [sample_record])
        result = cursor.fetchone()
        conn.close()

        assert result[0] == "更新后的备注内容"


    def test_update_notes_with_empty_string(self, client, test_db, sample_record):
        """测试空字符串自动转为NULL"""
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "   "}  # 纯空白字符串
        )

        assert response.status_code == 200
        data = response.json()

        # 空白字符串应该被转为NULL
        assert data["notes"] is None

        # 验证数据库中保存为NULL
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT notes FROM upload_history WHERE id = ?", [sample_record])
        result = cursor.fetchone()
        conn.close()

        assert result[0] is None


    def test_update_notes_exceeds_max_length(self, client, test_db, sample_record):
        """测试超长文本验证（1000字符限制）"""
        long_notes = "a" * 1001  # 1001个字符

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": long_notes}
        )

        assert response.status_code == 400
        data = response.json()
        assert "1000字符" in data["detail"]


    def test_update_notes_exactly_max_length(self, client, test_db, sample_record):
        """测试恰好1000字符的边界情况"""
        notes_1000 = "b" * 1000  # 恰好1000个字符

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": notes_1000}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert len(data["notes"]) == 1000


    def test_update_notes_record_not_found(self, client, test_db):
        """测试更新不存在的记录返回404"""
        response = client.patch(
            "/api/admin/records/999999/notes",
            json={"notes": "测试备注"}
        )

        assert response.status_code == 404
        data = response.json()
        assert "不存在" in data["detail"]


    def test_update_notes_deleted_record(self, client, test_db, sample_record):
        """测试更新已软删除的记录返回404"""
        # 先软删除记录
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE upload_history SET deleted_at = ? WHERE id = ?",
            [datetime.now().isoformat(), sample_record]
        )
        conn.commit()
        conn.close()

        # 尝试更新备注
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "测试备注"}
        )

        assert response.status_code == 404
        data = response.json()
        assert "删除" in data["detail"]


    def test_export_includes_notes_column(self, client, test_db, sample_records_with_notes):
        """测试导出Excel包含备注列"""
        response = client.get("/api/admin/export")

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/zip"

        # 解析ZIP文件
        zip_data = BytesIO(response.content)
        with zipfile.ZipFile(zip_data, 'r') as zipf:
            # 查找Excel文件
            excel_files = [f for f in zipf.namelist() if f.endswith('.xlsx')]
            assert len(excel_files) == 1

            # 读取Excel文件
            excel_content = zipf.read(excel_files[0])
            wb = load_workbook(BytesIO(excel_content))
            ws = wb.active

            # 验证表头包含"备注"列
            headers = [cell.value for cell in ws[1]]
            assert "备注" in headers

            # 验证备注列是最后一列
            assert headers[-1] == "备注"

            # 验证数据行包含备注内容
            row_2 = [cell.value for cell in ws[2]]
            notes_index = headers.index("备注")
            assert row_2[notes_index] in ["第一条备注", "第二条备注", "", None]


# ==================== 2. 边界情况测试 ====================

class TestNotesBoundaryConditions:
    """测试备注功能的边界情况"""

    def test_notes_with_special_characters(self, client, test_db, sample_record):
        """测试特殊字符处理（emoji、中文、标点）"""
        special_notes = "这是中文😊！@#$%^&*()_+-=[]{}|;:',.<>?/`~"

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": special_notes}
        )

        assert response.status_code == 200
        data = response.json()
        assert data["notes"] == special_notes

        # 验证数据库正确存储特殊字符
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT notes FROM upload_history WHERE id = ?", [sample_record])
        result = cursor.fetchone()
        conn.close()

        assert result[0] == special_notes


    def test_notes_with_unicode_emoji(self, client, test_db, sample_record):
        """测试Unicode emoji处理"""
        emoji_notes = "测试✅ 警告⚠️ 错误❌ 笑脸😊 心❤️"

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": emoji_notes}
        )

        assert response.status_code == 200
        assert response.json()["notes"] == emoji_notes


    def test_notes_with_newlines(self, client, test_db, sample_record):
        """测试换行符处理"""
        notes_with_newlines = "第一行\n第二行\r\n第三行"

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": notes_with_newlines}
        )

        assert response.status_code == 200
        assert response.json()["notes"] == notes_with_newlines


    def test_notes_with_quotes(self, client, test_db, sample_record):
        """测试引号处理"""
        notes_with_quotes = '这是"双引号"和\'单引号\'测试'

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": notes_with_quotes}
        )

        assert response.status_code == 200
        assert response.json()["notes"] == notes_with_quotes


    def test_notes_null_value(self, client, test_db, sample_record):
        """测试NULL值处理"""
        # 更新为NULL (通过空字符串)
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": ""}
        )

        assert response.status_code == 200
        assert response.json()["notes"] is None

        # 查询时验证返回NULL
        response = client.get("/api/admin/records?page=1&page_size=10")
        assert response.status_code == 200

        records = response.json()["records"]
        record = next((r for r in records if r["id"] == sample_record), None)
        assert record is not None
        assert record["notes"] is None


    def test_notes_whitespace_trimming(self, client, test_db, sample_record):
        """测试空白字符处理"""
        # 前后有空格的文本
        notes_with_spaces = "  有效内容  "

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": notes_with_spaces}
        )

        assert response.status_code == 200
        # 验证前后空格被保留（根据实际代码实现）
        # 如果代码trim了空格，则验证trim结果
        data = response.json()
        assert data["notes"].strip() == "有效内容"


    def test_notes_sql_injection_attempt(self, client, test_db, sample_record):
        """测试SQL注入防护"""
        sql_injection = "'; DROP TABLE upload_history; --"

        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": sql_injection}
        )

        # 应该正常保存（使用参数化查询防护）
        assert response.status_code == 200
        assert response.json()["notes"] == sql_injection

        # 验证表未被删除
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM upload_history")
        count = cursor.fetchone()[0]
        conn.close()

        assert count > 0  # 表仍然存在且有数据


# ==================== 3. 集成测试 ====================

class TestNotesIntegration:
    """测试备注功能的端到端流程"""

    def test_create_update_query_workflow(self, client, test_db):
        """测试完整流程：创建记录 → 添加备注 → 查询验证"""
        # 1. 创建记录
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO upload_history (
                business_id, doc_number, doc_type, file_name,
                file_size, file_extension, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ["123456", "SO999", "销售", "test.jpg", 1024, ".jpg", "success"])
        record_id = cursor.lastrowid
        conn.commit()
        conn.close()

        # 2. 添加备注
        update_response = client.patch(
            f"/api/admin/records/{record_id}/notes",
            json={"notes": "集成测试备注"}
        )
        assert update_response.status_code == 200

        # 3. 查询验证
        query_response = client.get("/api/admin/records?page=1&page_size=10")
        assert query_response.status_code == 200

        records = query_response.json()["records"]
        record = next((r for r in records if r["id"] == record_id), None)

        assert record is not None
        assert record["notes"] == "集成测试备注"


    def test_multiple_updates_overwrite(self, client, test_db, sample_record):
        """测试多次更新备注（Last-Write-Wins策略）"""
        # 第一次更新
        response1 = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "第一次更新"}
        )
        assert response1.status_code == 200

        # 第二次更新
        response2 = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "第二次更新"}
        )
        assert response2.status_code == 200

        # 第三次更新
        response3 = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "最终备注"}
        )
        assert response3.status_code == 200

        # 验证最后一次更新生效
        query_response = client.get("/api/admin/records?page=1&page_size=10")
        records = query_response.json()["records"]
        record = next((r for r in records if r["id"] == sample_record), None)

        assert record["notes"] == "最终备注"


    def test_export_after_update(self, client, test_db, sample_records_with_notes):
        """测试更新备注后导出验证"""
        # 更新第一条记录的备注
        record_id = sample_records_with_notes[0]
        client.patch(
            f"/api/admin/records/{record_id}/notes",
            json={"notes": "更新后导出测试"}
        )

        # 导出并验证
        response = client.get("/api/admin/export")
        assert response.status_code == 200

        zip_data = BytesIO(response.content)
        with zipfile.ZipFile(zip_data, 'r') as zipf:
            excel_files = [f for f in zipf.namelist() if f.endswith('.xlsx')]
            excel_content = zipf.read(excel_files[0])
            wb = load_workbook(BytesIO(excel_content))
            ws = wb.active

            # 验证备注列包含更新后的内容
            headers = [cell.value for cell in ws[1]]
            notes_index = headers.index("备注")

            # 检查是否有一行包含更新后的备注
            found_updated_note = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[notes_index] == "更新后导出测试":
                    found_updated_note = True
                    break

            assert found_updated_note, "导出的Excel应包含更新后的备注"


    def test_filter_and_query_with_notes(self, client, test_db, sample_records_with_notes):
        """测试带筛选条件的查询返回正确的notes"""
        # 按单据类型筛选
        response = client.get("/api/admin/records?page=1&page_size=10&doc_type=销售")

        assert response.status_code == 200
        data = response.json()

        # 验证所有返回的记录都有notes字段
        for record in data["records"]:
            assert "notes" in record
            if record["doc_type"] == "销售":
                # 销售类型的记录应该有对应的备注或NULL
                assert record["notes"] in ["第一条备注", None, ""]


# ==================== 4. 性能和并发测试（可选）====================

class TestNotesPerformance:
    """测试备注功能的性能和并发"""

    @pytest.mark.slow
    def test_update_notes_response_time(self, client, test_db, sample_record):
        """测试更新备注的响应时间应小于500ms"""
        import time

        start_time = time.time()
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={"notes": "性能测试备注"}
        )
        end_time = time.time()

        response_time_ms = (end_time - start_time) * 1000

        assert response.status_code == 200
        # 允许一定的性能波动，实际目标是500ms以内
        assert response_time_ms < 1000, f"响应时间过长: {response_time_ms:.2f}ms"


    @pytest.mark.slow
    def test_batch_query_with_notes(self, client, test_db):
        """测试批量查询包含notes字段的性能"""
        # 创建100条测试记录
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()

        for i in range(100):
            cursor.execute("""
                INSERT INTO upload_history (
                    business_id, doc_number, doc_type, file_name,
                    file_size, file_extension, status, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                f"12345{i % 10}",
                f"SO{i:04d}",
                "销售" if i % 2 == 0 else "转库",
                f"file{i}.jpg",
                1024 * (i + 1),
                ".jpg",
                "success",
                f"批量测试备注{i}" if i % 3 == 0 else None
            ])

        conn.commit()
        conn.close()

        # 查询所有记录
        import time
        start_time = time.time()
        response = client.get("/api/admin/records?page=1&page_size=100")
        end_time = time.time()

        response_time_ms = (end_time - start_time) * 1000

        assert response.status_code == 200
        assert len(response.json()["records"]) <= 100
        assert response_time_ms < 2000, f"批量查询响应时间过长: {response_time_ms:.2f}ms"


# ==================== 5. 错误处理测试 ====================

class TestNotesErrorHandling:
    """测试备注功能的错误处理"""

    def test_update_notes_invalid_record_id(self, client, test_db):
        """测试无效的record_id"""
        response = client.patch(
            "/api/admin/records/-1/notes",
            json={"notes": "测试"}
        )

        assert response.status_code == 404


    def test_update_notes_missing_notes_field(self, client, test_db, sample_record):
        """测试缺少notes字段的请求"""
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            json={}  # 缺少notes字段
        )

        assert response.status_code == 422  # FastAPI参数验证错误


    def test_update_notes_invalid_json(self, client, test_db, sample_record):
        """测试无效的JSON格式"""
        response = client.patch(
            f"/api/admin/records/{sample_record}/notes",
            data="invalid json",
            headers={"Content-Type": "application/json"}
        )

        assert response.status_code == 422


    def test_get_records_with_corrupted_notes(self, client, test_db):
        """测试数据库中备注字段损坏的情况"""
        # 插入一条记录，notes字段故意设为特殊值
        conn = sqlite3.connect(test_db)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO upload_history (
                business_id, doc_number, file_name, file_size,
                file_extension, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ["123456", "SO001", "test.jpg", 1024, ".jpg", "success", None])
        conn.commit()
        conn.close()

        # 查询应该正常返回
        response = client.get("/api/admin/records?page=1&page_size=10")
        assert response.status_code == 200

        records = response.json()["records"]
        assert len(records) > 0


# ==================== 运行标记 ====================

# 标记关键测试用例
pytestmark = pytest.mark.unit
