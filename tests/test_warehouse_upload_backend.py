import os
import sqlite3
from contextlib import contextmanager
from io import BytesIO
from unittest.mock import AsyncMock, Mock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.core.timezone import get_beijing_now_naive
from app.api.upload import background_save_warehouse_upload


client = TestClient(app)


@contextmanager
def db_context(db_path):
    conn = sqlite3.connect(db_path)
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def patch_db(module_name, db_path):
    return patch(f"{module_name}.get_db_connection", side_effect=lambda: db_context(db_path))


def seed_upload_record(
    db_path,
    *,
    business_id="123456",
    doc_number="SO001",
    upload_type=None,
    status="success",
    deleted_at=None,
    yonyou_file_id=None,
    logistics=None,
    file_name=None,
):
    now = get_beijing_now_naive().isoformat()
    file_name = file_name or f"{doc_number}.jpg"
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO upload_history
            (business_id, doc_number, doc_type, product_type, upload_type, file_name,
             file_size, file_extension, upload_time, status, yonyou_file_id, logistics,
             retry_count, local_file_path, created_at, updated_at, deleted_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                business_id,
                doc_number,
                "销售",
                "油脂",
                upload_type,
                file_name,
                128,
                ".jpg",
                now,
                status,
                yonyou_file_id,
                logistics,
                0,
                None,
                now,
                now,
                deleted_at,
            ),
        )
        return cursor.lastrowid


class TestWarehouseUploadApi:
    def test_upload_defaults_missing_type_to_logistics(self, test_db_path, test_image_bytes):
        with patch_db("app.api.upload", test_db_path), \
             patch("app.api.upload.background_upload_to_yonyou", new_callable=AsyncMock) as logistics_task:
            response = client.post(
                "/api/upload",
                files=[("files", ("test.jpg", test_image_bytes, "image/jpeg"))],
                data={"business_id": "123456", "doc_number": "SO001", "doc_type": "销售"},
            )

        assert response.status_code == 200
        record = response.json()["records"][0]
        assert record["upload_type"] == "物流"
        assert logistics_task.called

        with sqlite3.connect(test_db_path) as conn:
            upload_type = conn.execute("SELECT upload_type FROM upload_history").fetchone()[0]
        assert upload_type == "物流"

    def test_upload_rejects_invalid_type(self, test_db_path, test_image_bytes):
        with patch_db("app.api.upload", test_db_path):
            response = client.post(
                "/api/upload",
                files=[("files", ("test.jpg", test_image_bytes, "image/jpeg"))],
                data={
                    "business_id": "123456",
                    "doc_number": "SO001",
                    "doc_type": "销售",
                    "upload_type": "门店",
                },
            )

        assert response.status_code == 400
        assert "upload_type" in response.json()["detail"]

    def test_logistics_type_schedules_yonyou_task(self, test_db_path, test_image_bytes):
        with patch_db("app.api.upload", test_db_path), \
             patch("app.api.upload.background_upload_to_yonyou", new_callable=AsyncMock) as logistics_task, \
             patch("app.api.upload.background_save_warehouse_upload", new_callable=AsyncMock) as warehouse_task:
            response = client.post(
                "/api/upload",
                files=[("files", ("test.jpg", test_image_bytes, "image/jpeg"))],
                data={
                    "business_id": "123456",
                    "doc_number": "SO001",
                    "doc_type": "销售",
                    "upload_type": "物流",
                },
            )

        assert response.status_code == 200
        assert logistics_task.called
        assert not warehouse_task.called

    def test_warehouse_upload_persists_type_and_schedules_storage_task(self, test_db_path, test_image_bytes):
        with patch_db("app.api.upload", test_db_path), \
             patch("app.api.upload.background_upload_to_yonyou", new_callable=AsyncMock) as logistics_task, \
             patch("app.api.upload.background_save_warehouse_upload", new_callable=AsyncMock) as warehouse_task:
            response = client.post(
                "/api/upload",
                files=[("files", ("warehouse.jpg", test_image_bytes, "image/jpeg"))],
                data={
                    "business_id": "123456",
                    "doc_number": "WH001",
                    "doc_type": "销售",
                    "upload_type": "仓库",
                },
            )

        assert response.status_code == 200
        assert not logistics_task.called
        assert warehouse_task.called
        assert response.json()["records"][0]["upload_type"] == "仓库"

        with sqlite3.connect(test_db_path) as conn:
            upload_type = conn.execute("SELECT upload_type FROM upload_history").fetchone()[0]
        assert upload_type == "仓库"


class TestWarehouseBackgroundTask:
    @pytest.mark.asyncio
    async def test_warehouse_save_success_uses_file_manager_and_not_yonyou(self, test_db_path, test_image_bytes, tmp_path, monkeypatch):
        monkeypatch.setattr("app.core.database.settings.DATABASE_URL", f"sqlite:///{test_db_path}")
        record_id = seed_upload_record(test_db_path, upload_type="仓库", status="pending")

        save_result = {
            "success": True,
            "webdav_path": "files/2026/06/03/warehouse.jpg",
            "local_cache_path": str(tmp_path / "warehouse.jpg"),
            "upload_time": get_beijing_now_naive().isoformat(),
            "file_size": len(test_image_bytes),
            "is_cached": True,
            "webdav_etag": "etag-1",
            "is_synced": True,
        }

        with patch("app.api.upload.file_manager.save_file", new_callable=AsyncMock, return_value=save_result) as save_file, \
             patch("app.api.upload.yonyou_client.upload_file", new_callable=AsyncMock) as upload_file, \
             patch("app.api.upload.yonyou_client.get_delivery_detail", new_callable=AsyncMock) as get_detail:
            await background_save_warehouse_upload(
                test_image_bytes,
                "warehouse.jpg",
                str(tmp_path / "fallback.jpg"),
                record_id,
            )

        save_file.assert_awaited_once()
        upload_file.assert_not_called()
        get_detail.assert_not_called()

        with sqlite3.connect(test_db_path) as conn:
            row = conn.execute(
                """
                SELECT status, yonyou_file_id, logistics, error_code, error_message,
                       webdav_path, is_cached, cache_expiry_time, retry_count
                FROM upload_history WHERE id = ?
                """,
                (record_id,),
            ).fetchone()
            metadata_count = conn.execute("SELECT COUNT(*) FROM file_metadata").fetchone()[0]

        assert row[0] == "success"
        assert row[1] is None
        assert row[2] is None
        assert row[3] is None
        assert row[4] is None
        assert row[5] == save_result["webdav_path"]
        assert row[6] == 1
        assert row[7] is not None
        assert row[8] == 0
        assert metadata_count == 1

    @pytest.mark.asyncio
    async def test_warehouse_save_failure_sets_storage_error(self, test_db_path, test_image_bytes, tmp_path, monkeypatch):
        monkeypatch.setattr("app.core.database.settings.DATABASE_URL", f"sqlite:///{test_db_path}")
        record_id = seed_upload_record(test_db_path, upload_type="仓库", status="pending")

        with patch("app.api.upload.file_manager.save_file", new_callable=AsyncMock, return_value={"success": False, "error": "webdav down"}), \
             patch("app.api.upload.save_file_locally", side_effect=OSError("disk full")), \
             patch("app.api.upload.yonyou_client.upload_file", new_callable=AsyncMock) as upload_file, \
             patch("app.api.upload.yonyou_client.get_delivery_detail", new_callable=AsyncMock) as get_detail:
            await background_save_warehouse_upload(
                test_image_bytes,
                "warehouse.jpg",
                str(tmp_path / "fallback.jpg"),
                record_id,
            )

        upload_file.assert_not_called()
        get_detail.assert_not_called()

        with sqlite3.connect(test_db_path) as conn:
            row = conn.execute(
                "SELECT status, error_code, error_message, yonyou_file_id, logistics FROM upload_history WHERE id = ?",
                (record_id,),
            ).fetchone()

        assert row[0] == "failed"
        assert row[1] == "WAREHOUSE_STORAGE_ERROR"
        assert "disk full" in row[2]
        assert row[3] is None
        assert row[4] is None


class TestWarehouseHistoryAdminExport:
    def test_history_filters_type_and_counts_legacy_null_as_logistics(self, test_db_path):
        seed_upload_record(test_db_path, business_id="900001", doc_number="LEGACY", upload_type=None)
        seed_upload_record(test_db_path, business_id="900001", doc_number="LOGI", upload_type="物流")
        seed_upload_record(test_db_path, business_id="900001", doc_number="WARE", upload_type="仓库")
        seed_upload_record(test_db_path, business_id="900001", doc_number="DELETED", upload_type="仓库", deleted_at=get_beijing_now_naive().isoformat())
        seed_upload_record(test_db_path, business_id="900001", doc_number="DELLOGI", upload_type="物流", deleted_at=get_beijing_now_naive().isoformat())

        with patch_db("app.api.history", test_db_path):
            logistics_response = client.get("/api/history/900001?upload_type=物流")
            warehouse_response = client.get("/api/history/900001?upload_type=仓库")
            all_response = client.get("/api/history/900001")

        assert logistics_response.status_code == 200
        logistics_records = logistics_response.json()["records"]
        logistics_file_names = {r["file_name"] for r in logistics_records}
        assert logistics_file_names == {"LEGACY.jpg", "LOGI.jpg"}
        assert "DELLOGI.jpg" not in logistics_file_names
        assert {r["upload_type"] for r in logistics_records} == {"物流"}

        assert warehouse_response.status_code == 200
        assert [r["file_name"] for r in warehouse_response.json()["records"]] == ["WARE.jpg"]
        assert warehouse_response.json()["records"][0]["upload_type"] == "仓库"

        assert all_response.status_code == 200
        assert all_response.json()["total_count"] == 3

    def test_history_rejects_invalid_upload_type(self, test_db_path):
        with patch_db("app.api.history", test_db_path):
            response = client.get("/api/history/900001?upload_type=门店")

        assert response.status_code == 400

    def test_admin_records_filter_by_upload_type(self, test_db_path):
        seed_upload_record(test_db_path, doc_number="LEGACY", upload_type=None)
        seed_upload_record(test_db_path, doc_number="LOGI", upload_type="物流")
        seed_upload_record(test_db_path, doc_number="WARE", upload_type="仓库")

        with patch_db("app.api.admin", test_db_path):
            logistics_response = client.get("/api/admin/records?upload_type=物流&page_size=20")
            warehouse_response = client.get("/api/admin/records?upload_type=仓库&page_size=20")
            invalid_response = client.get("/api/admin/records?upload_type=门店")

        assert logistics_response.status_code == 200
        assert {r["doc_number"] for r in logistics_response.json()["records"]} == {"LEGACY", "LOGI"}
        assert {r["upload_type"] for r in logistics_response.json()["records"]} == {"物流"}

        assert warehouse_response.status_code == 200
        assert [r["doc_number"] for r in warehouse_response.json()["records"]] == ["WARE"]
        assert warehouse_response.json()["records"][0]["upload_type"] == "仓库"

        assert invalid_response.status_code == 400

    def test_export_filter_and_header_upload_type_first(self, test_db_path):
        seed_upload_record(test_db_path, doc_number="LOGI", upload_type="物流")
        seed_upload_record(test_db_path, doc_number="WARE", upload_type="仓库")

        with patch_db("app.api.admin", test_db_path):
            response = client.get("/api/admin/export?upload_type=仓库&include_images=false")

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == ["上传业务类型", "单据编号", "单据类型", "产品类型", "业务ID", "上传时间", "文件名", "文件大小(字节)", "状态", "备注"]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][0] == "仓库"
        assert rows[0][1] == "WARE"

    def test_statistics_includes_by_upload_type(self, test_db_path):
        seed_upload_record(test_db_path, doc_number="LEGACY", upload_type=None)
        seed_upload_record(test_db_path, doc_number="BLANK", upload_type="")
        seed_upload_record(test_db_path, doc_number="LOGI", upload_type="物流")
        seed_upload_record(test_db_path, doc_number="WARE", upload_type="仓库")
        seed_upload_record(test_db_path, doc_number="DELETED", upload_type="仓库", deleted_at=get_beijing_now_naive().isoformat())

        with patch_db("app.api.admin", test_db_path):
            response = client.get("/api/admin/statistics")

        assert response.status_code == 200
        by_upload_type = response.json()["by_upload_type"]
        assert by_upload_type["物流"] == 3
        assert by_upload_type["仓库"] == 1


class TestWarehouseSchemaAndModel:
    def test_test_schema_has_upload_type_and_webdav_columns(self, test_db_path):
        with sqlite3.connect(test_db_path) as conn:
            columns = {row[1] for row in conn.execute("PRAGMA table_info(upload_history)").fetchall()}
            indexes = {row[1] for row in conn.execute("PRAGMA index_list(upload_history)").fetchall()}

        assert {"upload_type", "webdav_path", "is_cached", "cache_expiry_time"}.issubset(columns)
        assert "idx_upload_type" in indexes

    def test_upload_history_model_defaults_to_logistics(self):
        from app.models.upload_history import UploadHistory

        history = UploadHistory()
        assert history.upload_type == "物流"
